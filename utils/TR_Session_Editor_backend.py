import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import re
from io import BytesIO


def read_woorkbook(upload_file): 
    return load_workbook(filename=BytesIO(upload_file.read()))


def process_workbook(workbook):
    sheet_TR_Matrix = workbook['TR Matrix']

    # Apply font changes to columns C, D, E (columns 3, 4, 5)
    for col_idx in [3, 4, 5]:
        col_letter = get_column_letter(col_idx)
        for row in range(2, sheet_TR_Matrix.max_row + 1):  # Skip header
            cell = sheet_TR_Matrix[f"{col_letter}{row}"]
            original_font = cell.font
            cell.font = Font(
                name="Calibri",
                color="FFFFFF",
                bold = original_font.bold,
                size=original_font.size if original_font and original_font.size else 11  # Default to 11 if missing
            )

    # Remove gridlines
    sheet_TR_Matrix.sheet_view.showGridLines = False


    #-----------------------------------------------------


    sheet_list = workbook['List View']

    columns_to_delete = ['J', 'I', 'H', 'G', 'D', 'B']
    for col_letter in columns_to_delete:
        col_idx = sheet_list[col_letter + '1'].column
        sheet_list.delete_cols(col_idx)

    # === 2. Set column E width to 120 (after deletion it's still E) ===
    sheet_list.column_dimensions['F'].width = 120

    # === 3. Center align all cells (preserve font/fill/etc.) ===
    for row in sheet_list.iter_rows(min_row=1, max_row=sheet_list.max_row, max_col=sheet_list.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text= True)

    # === 4. Clean column E: remove everything after 'Created by' (case-insensitive) ===
    for row in range(2, sheet_list.max_row + 1):
        cell = sheet_list[f'F{row}']
        if cell.value and isinstance(cell.value, str):
            match = re.split(r'Created by', cell.value, flags=re.IGNORECASE)
            if match:
                cell.value = match[0].strip()

    sheet_list.sheet_view.showGridLines = False

    sheet_list.freeze_panes = "B1"

    return workbook

def save_workbook_to_bytes(workbook):
    # Save the workbook to a bytes buffer and return it
    bio = BytesIO()
    workbook.save(bio)
    bio.seek(0)
    return bio

